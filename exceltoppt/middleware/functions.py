import openpyxl
import os
import openai
import json
from openai import OpenAI

from fastapi import UploadFile, File
from dotenv import load_dotenv
import aspose.slides as slides

import openpyxl
import pandas as pd


load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")
os.environ["OPENAI_API_KEY"] = os.getenv("OPENAI_API_KEY")
current_woring_dir = os.path.dirname(os.path.realpath(__file__))

def convert_bytes_to_human_readable(size_in_bytes):
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size_in_bytes < 1024.0:
            return f"{size_in_bytes:.2f} {unit}"
        size_in_bytes /= 1024.0

async def upload_file(file: UploadFile = File(...)):
    try:

        #Get File Type
        file_type = file.content_type

        #Get File Size
        file.file.seek(0,2)
        size = file.file.tell()
        file.file.seek(0)
        file_size = convert_bytes_to_human_readable(size)

        #Get File Name
        file_name = file.filename

        #Save Excel file to local drive
        os.makedirs("data/Excel", exist_ok=True)
        with open(f"data/Excel/{file.filename}", "wb") as f:
            f.write(file.file.read())

        #Excel/Workbook - get list of sheets
        sheets = []
        sheets_count = 0
        if file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or file_type == "application/vnd.ms-excel":
            file_path = f"data/Excel/{file.filename}"
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheets = workbook.sheetnames
            sheets_count = len(sheets)
        return {"response": "Success!", "file_size": file_size, "file_name": file_name, "file_type" : file_type, "sheet": sheets, "sheets_count": sheets_count}
    except Exception as e:
        return {"response": "Error!", "error_message": e}


def api_openai(content: str):
    client = OpenAI()
    chat_completion = client.chat.completions.create(
        messages=[
            {"role":"system", "content":"You are a helpful assistant."},
            {"role":"user","content":content}

        ],
        model="gpt-4",
    )
    return(chat_completion.choices[0].message.content)



#workbook_file_path = "/Users/suchi_bigmac/Documents/elevaite/exceltoppt/middleware/data/Excel/GAAP_RECON_NOTE_1.xlsx"
#workbook_file_path = "/Users/suchi_bigmac/Documents/elevaite/exceltoppt/middleware/data/Excel/cisco.xlsx"
workbook_file_path = "/Users/suchi_bigmac/Documents/elevaite/exceltoppt/middleware/data/Excel/GAPP_RECON_NOTE_2.xlsx"
import openpyxl

def find_numeric_cells(workbook_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    # Select the sheet by name
    sheet = workbook.active

    # Initialize variables to store the first and last row numbers and column numbers
    first_row = None
    last_row = None
    first_col = None
    last_col = None

    # Iterate through rows and columns in the sheet
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_index, cell_value in enumerate(row, start=1):
            # Check if the cell value is numeric
            if isinstance(cell_value, (int, float)):
                # Update first_row if not set
                if first_row is None:
                    first_row = row_index

                # Update last_row continuously
                last_row = row_index

                # Update first_col if not set
                if first_col is None:
                    first_col = col_index

                # Update last_col continuously
                if last_col is None:
                    last_col = col_index
                if last_col < col_index:
                    last_col = col_index

    return first_row, first_col, last_row, last_col

def create_dataframe(start_row, start_col, end_row, end_col, row_header, col_header, workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    # Select the sheet by name
    sheet = workbook.active
    
    rows_data = []

    # Iterate through rows and columns in the specified range and populate the list
    for row_index in range(start_row, end_row + 1):
        row_data = {}
        for col_index in range(start_col, end_col + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            row_data[f'Column {col_index}'] = cell_value

        rows_data.append(row_data)

    # Create DataFrame from the list of dictionaries
    df = pd.DataFrame(rows_data)
    df.index = row_header
    df.columns = col_header
    
    csv_file = os.path.basename(workbook_file_path).split(".")[0] + ".csv"
    csv_file_path = os.path.join(current_woring_dir, "data", "Output", csv_file)
    if not os.path.exists(csv_file_path):
        mode = 'w'
        df.to_csv(csv_file_path, mode = mode)
    
    return df

def get_index(start_row, end_row, start_column, end_column, workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook.active

    row_headers = []
    col_headers = []

    for row_index in range(start_row, end_row + 1):
        header = ""
        for col_index in range(0, start_column):
            try:
                cell_value = sheet.cell(row=row_index, column=col_index).value
                if header == "" and cell_value is not None:
                    header = str(cell_value)
                elif header != "" and cell_value is not None:
                    header+=str(cell_value)
            except Exception as e:
                print("..")
        
        row_headers.append(header)
    for col_index in range(start_column, end_column+1):
        header = ""
        for row_index in range(0, start_row):
            try:
                cell_value = sheet.cell(row=row_index, column=col_index).value
                
                if header == "" and cell_value is not None:
                    header = str(cell_value)
                elif header != "" and cell_value is not None:
                    header += "_"+str(cell_value)
            except Exception as e:
                print("..")
        col_headers.append(header)
    return row_headers, col_headers

